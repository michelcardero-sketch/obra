"""Regenerates data.js and data-servicos.js from the source Excel workbook.

Run this after editing "Controle de Gastos.xlsx", before publishing. It only
rewrites the two data files here in "BI em HTML" -- it never touches the
Excel file itself.
"""
import json
import os
import sys

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(os.path.dirname(HERE), "Controle de Gastos.xlsx")


def load_sheet(wb, name):
    if name not in wb.sheetnames:
        sys.exit(f"Aba '{name}' nao encontrada no Excel. Abas disponiveis: {wb.sheetnames}")
    return wb[name]


def export_custo_total(wb):
    ws = load_sheet(wb, "Custo Total")
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    records = []
    for data, etapa, descricao, fornecedor, sacado, valor, tipo, *_ in rows:
        if data is None:
            continue
        records.append({
            "data": data.strftime("%Y-%m-%d"),
            "etapa": str(etapa).strip(),
            "descricao": str(descricao).strip(),
            "fornecedor": str(fornecedor).strip(),
            "sacado": str(sacado).strip(),
            "valor": round(float(valor), 2),
            "tipo": str(tipo).strip(),
        })

    out_path = os.path.join(HERE, "data.js")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("const RAW_DATA = ")
        json.dump(records, f, ensure_ascii=False, indent=2)
        f.write(";\n")
    print(f"data.js: {len(records)} lancamentos")


def export_servicos(wb):
    ws = load_sheet(wb, "Contrato Serviços")
    contrato = []
    for data, etapa, descricao, sacado, valor, entregue in ws.iter_rows(min_row=2, values_only=True):
        if etapa is None:
            continue
        contrato.append({
            "etapa": str(etapa).strip(),
            "descricao": str(descricao).strip() if descricao else "",
            "valor": round(float(valor), 2) if valor is not None else None,
            "entregue": bool(entregue and str(entregue).strip().upper() == "SIM"),
        })

    ws2 = load_sheet(wb, "Pagamentos Serviços")
    pagamentos = []
    for data, etapa, descricao, sacado, valor in ws2.iter_rows(min_row=2, values_only=True):
        if data is None:
            continue
        pagamentos.append({
            "data": data.strftime("%Y-%m-%d"),
            "etapa": str(etapa).strip(),
            "descricao": str(descricao).strip(),
            "sacado": str(sacado).strip(),
            "valor": round(float(valor), 2),
        })

    out_path = os.path.join(HERE, "data-servicos.js")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("const CONTRATO_DATA = ")
        json.dump(contrato, f, ensure_ascii=False, indent=2)
        f.write(";\nconst PAGAMENTOS_DATA = ")
        json.dump(pagamentos, f, ensure_ascii=False, indent=2)
        f.write(";\n")
    print(f"data-servicos.js: {len(contrato)} etapas de contrato, {len(pagamentos)} pagamentos")


def main():
    if not os.path.exists(SRC):
        sys.exit(f"Excel nao encontrado em: {SRC}")
    wb = openpyxl.load_workbook(SRC, data_only=True)
    export_custo_total(wb)
    export_servicos(wb)


if __name__ == "__main__":
    main()
